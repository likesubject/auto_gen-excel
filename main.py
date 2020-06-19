# -- coding: utf-8 --

import calendar
import contextlib
import datetime
import json
import logging
import os
from collections import OrderedDict
from typing import Generator

import click
import requests
from jinja2 import Template
from lxml import html
from openpyxl import load_workbook
from redminelib import Redmine
from win32com.client import Dispatch


logger = logging.getLogger(__name__)
etree = html.etree


class LocalResourceBase(object):
    include_attributes = [
    ]

    redmine_resource = None

    #  {attribute: redmine_custom_attribute}
    custom_attributes = {
        'custom': '',
    }

    def __init__(self, name: str, uid: int, resources: [], **kwargs):
        self.name = name
        self.uid = uid
        self._resources = OrderedDict()
        self.extend_resource(resources)
        self._remote_resource = kwargs.pop('remote_resource', None)
        self._redmine = kwargs.pop('redmine', None)
        self._cached_items = {}
        self._cached_remote_resource = None

    def append_resource(self, resource):
        self._resources.update({resource.uid: resource})

    def extend_resource(self, resources):
        for resource in resources:
            self.append_resource(resource)

    def clear_resource(self):
        self._resources = {}

    @property
    def resources(self):
        for key, val in self._resources.items():
            yield val

    def get_resource_by_uid(self, uid: int):
        return self._resources.get(uid)

    def get_resource(self, remote_resource, cls):
        resource = self.get_resource_by_uid(remote_resource.id)
        if resource is None:
            #  name is optional
            resource = cls(getattr(remote_resource, 'name', None),
                           remote_resource.id, [], remote_resource=remote_resource, redmine=self._redmine)
            self.append_resource(resource)
        return resource

    def cache_data(self, attribute, data):
        self._cached_items.update({attribute: data})

    def get_cached_data(self, attribute):
        return self._cached_items.get(attribute, None)

    def __getattr__(self, item):
        value = getattr(self._remote_resource, item, None)
        if value is None:
            value = getattr(self.cached_remote_resource, item, None)
        if value is None:
            value = self.get_custom_attributes(item)
        return value

    @property
    def cached_remote_resource(self):
        if self._cached_remote_resource is None:
            resource = getattr(self._redmine, self.redmine_resource, None)
            if resource is not None:
                self._cached_remote_resource = resource.get(self.uid)
        return self._cached_remote_resource

    def get_custom_attributes(self, attribute):
        redmine_custom_attribute = self.custom_attributes.get(attribute, None)
        if redmine_custom_attribute is None or redmine_custom_attribute == '':
            return None
        value = self.get_cached_data(redmine_custom_attribute)
        if value is None:
            for fields in self.cached_remote_resource.custom_fields:
                if getattr(fields, 'name', None) == redmine_custom_attribute:
                    value = getattr(fields, 'value', None)
                    if value is not None:
                        self.cache_data(redmine_custom_attribute, value)
                        break
        return value


class WorkTime(LocalResourceBase):
    redmine_resource = 'time_entry'


class Task(LocalResourceBase):
    redmine_resource = 'issue'

    include_attributes = [
        'progress',
        'status',
        'start_time',
        'end_time',
    ]

    @property
    def work_times(self):
        return self.resources

    def get_work_time(self, remote_work_time) -> WorkTime:
        return super(Task, self).get_resource(remote_work_time, WorkTime)

    @property
    def spent_time(self) -> float:
        spent_time = self.get_cached_data('spent_time')
        if spent_time is None:
            spent_time = 0
            for work_time in self.work_times:
                hours = work_time.hours
                if isinstance(hours, float):
                    spent_time += hours
            self.cache_data('spent_time', spent_time)
        return spent_time


class User(LocalResourceBase):
    redmine_resource = 'user'

    @property
    def tasks(self):
        return self.resources

    def get_task(self, remote_task) -> Task:
        return super(User, self).get_resource(remote_task, Task)

    @property
    def spent_time(self):
        spent_time = self.get_cached_data('spent_time')
        if spent_time is None:
            spent_time = 0
            for task in self.tasks:
                spent_time += task.spent_time
            spent_time = round(spent_time / 8.0, 2)
            self.cache_data('spent_time', spent_time)
        return spent_time

    @property
    def fullname(self):
        return '{0}{1}'.format(self.lastname, self.firstname)


class Project(LocalResourceBase):
    redmine_resource = 'project'

    custom_attributes = {
        'custom_num': '项目编号',
        'custom_name': '项目名称',
        'custom_category': '产品分类',
        'custom_leader': '项目负责人',
        'custom_time': '立项时间',
    }

    @property
    def users(self) -> Generator[None, User, None]:
        return self.resources

    def get_user(self, remote_user) -> User:
        """
        get user instance
        :param remote_user: remote user instance
        :return: User instance
        """
        return super(Project, self).get_resource(remote_user, User)

    @property
    def parent_id(self):
        return getattr(self.parent, 'id', None)


class Users(LocalResourceBase):
    def __init__(self, redmine):
        super(Users, self).__init__(None, None, [], redmine=redmine)

    @property
    def users(self):
        return self.resources

    def get_user(self, remote_user) -> User:
        """
        get user instance
        :param remote_user: remote user instance
        :return: User instance
        """
        return super(Users, self).get_resource(remote_user, User)


class Projects(LocalResourceBase):
    def __init__(self, redmine):
        super(Projects, self).__init__(None, None, [], redmine=redmine)

    @property
    def projects(self) -> Generator[None, Project, None]:
        return self.resources

    def get_project(self, remote_project) -> Project:
        """
        get user instance
        :param remote_project: remote project instance
        :return: Project instance
        """
        return super(Projects, self).get_resource(remote_project, Project)

    def get_project_by_project_id(self, project_id):
        return super(Projects, self).get_resource_by_uid(project_id)


class CustomCell(object):
    def __init__(self, adapter, row_index, column_index):
        self.row_index = row_index
        self.column_index = column_index
        self.adapter = adapter
        self._cached_text = None

    def merge(self, custom_cell):
        self.adapter.current_workbook.\
            merge_cells(start_row=self.row_index,
                        start_column=self.column_index,
                        end_row=custom_cell.row_index,
                        end_column=custom_cell.column_index)

    def get_text(self):
        if self._cached_text is None:
            self._cached_text = self.adapter.get_text(self.row_index, self.column_index)
        return self._cached_text


class ExcelAdapter(object):
    def __init__(self, source_name: str, target_name: str):
        source_file_path = os.path.join(os.getcwd(), source_name)
        target_path = os.path.join(os.getcwd(), 'work tables')
        target_file_path = os.path.join(target_path, target_name)
        if not os.path.exists(target_path):
            os.mkdir(target_path)
        self.source_file_path = source_file_path
        self.target_file_path = target_file_path
        self.current_workbook = None
        self.error_flag = False
        self.source_name = source_name
        self.target_name = target_name

    def set_text(self, text, row_index=1, column_index=1):
        self.current_workbook.cell(column=column_index, row=row_index, value=text)

    def get_text(self, row_index=1, column_index=1):
        return self.current_workbook.cell(column=column_index, row=row_index).value

    def get_cell(self, row_index=1, column_index=1, **kwargs):
        return CustomCell(self, row_index, column_index)

    def get_cells(self, row_index=1):
        for column_index in range(1, 100):
            text = self.get_text(row_index=row_index, column_index=column_index)
            if text is not None and text != '':
                yield CustomCell(self, row_index, column_index)

    @staticmethod
    def merge(src_cell: CustomCell, dst_cell: CustomCell):
        src_cell.merge(dst_cell)

    @contextlib.contextmanager
    def context(self):
        template_workbook = load_workbook(filename=self.source_file_path)
        self.current_workbook = template_workbook.active
        yield self
        if not self.error_flag:
            template_workbook.save(self.target_file_path)

    def set_error_flag(self):
        self.error_flag = True

    def open_excel_for_windows(self):
        excel_app = Dispatch('Excel.Application')
        excel_app.Visible = 1
        excel_app.Workbooks.Open(self.target_file_path)


class WorkTable(object):
    def __init__(self, adapter: ExcelAdapter, projects: [],
                 start_row: int = 2, start_column: int = 1, enable_merge=True):
        self.adapter = adapter
        self.start_row = start_row
        self.start_column = start_column
        self.column_count = 0
        self.row_count = 0
        self._columns = []
        self._rows = self.pre_process(projects)
        self._cached_data = []
        self.enable_merge = enable_merge

    def parse(self):
        columns = []
        for column_index, first_row_cell in enumerate(self.adapter.get_cells(row_index=1)):
            second_row_cell = self.adapter.get_cell(column_index=column_index + 1, row_index=2)
            columns.append(ColumnRawData(column_index + 1,
                                         first_row_cell.get_text(),
                                         second_row_cell.get_text(),
                                         self.adapter))
        return columns

    def render(self, *args, **context):
        self.column_count = 0
        for column in self._columns:
            can_merge = False
            if column.can_render():
                if column.can_merge() and self.enable_merge:
                    can_merge = True
                data = column.render(*args, **context)
            else:
                data = column.get_text()
            self.write(can_merge, data)
            self.column_count += 1
        self.row_count += 1

    @property
    def current_column(self):
        return self.start_column + self.column_count

    @property
    def current_row(self):
        return self.start_row + self.row_count

    def _has_equal_cached_data(self, data):
        if not self._cached_data[self.column_count]:
            return False

        for key, value in self._cached_data[self.column_count].items():
            if value[1] != data:
                return False

        return True

    def merge_cur_column_cells(self):
        current_data = self._cached_data[self.column_count]
        self.merge_cells(current_data, self.column_count)

    def merge_cells(self, current_data, column_index):
        columns = list(current_data.items())
        if current_data and len(columns) >= 2:
            current_column = columns[0][1][0]
            current_row = columns[0][0]
            src_cell = self.adapter.get_cell(current_row, current_column)
            dst_cell = self.adapter.get_cell(columns[-1][0], current_column)
            self.adapter.set_text(columns[0][1][1], row_index=current_row, column_index=current_column)
            self.adapter.merge(src_cell, dst_cell)
            self._cached_data[column_index] = OrderedDict()
        elif current_data and len(columns) == 1:
            current_column = columns[0][1][0]
            current_row = columns[0][0]
            self.adapter.set_text(columns[0][1][1], row_index=current_row, column_index=current_column)
            self._cached_data[column_index] = OrderedDict()

    def merge_all_cells(self):
        for i, data in enumerate(self._cached_data):
            self.merge_cells(data, i)

    def write(self, can_merge: bool, data: str, **context):
        if can_merge:
            if not self._has_equal_cached_data(data):
                self.merge_cur_column_cells()
            self._cached_data[self.column_count].update(
                {self.current_row: (self.current_column, data)}
            )
        else:
            self.adapter.set_text(data, row_index=self.current_row, column_index=self.current_column)

    @staticmethod
    def pre_process(projects: [Project]):
        for project in projects.projects:
            for user in project.users:
                yield (project, user)

    def process(self):
        error_flag = True
        click.echo('Step three: Generating Excel,please waiting....')
        with self.adapter.context():
            self._columns = self.parse()
            self._cached_data.extend([OrderedDict() for i in self._columns])
            try:
                with click.progressbar(self._rows) as bar:
                    for project, user in bar:
                        error_flag = False
                        self.render(project=project, current_user=user)
                self.merge_all_cells()
                if error_flag:
                    self.adapter.set_error_flag()
                    raise ValueError('Unsuccessfully generated , no data was generated!!!')
            except Exception as e:
                self.adapter.set_error_flag()
                click.echo(str(e))
        if error_flag:
            raise Exception
        else:
            click.echo('Successfully generated, please open `{0}` file under the `work table` dir'.
                       format(self.adapter.target_name))
            self.adapter.open_excel_for_windows()


class ColumnRawData(object):
    def __init__(self, column_id, field_text, render_text, adapter):
        self.column_id = column_id
        self.field_text = field_text
        self.render_text = render_text
        self.template = Template(render_text)
        self.adapter = adapter

    def can_render(self):
        return '{{' in self.render_text

    def can_merge(self):
        return 'merge' in self.render_text

    def render(self, *args, **context):
        return self.template.render(*args, **context)

    def get_text(self):
        return self.render_text


class CustomRemoteProject(object):
    def __init__(self, id, name):
        self.id = id
        self.name = name


class RedmineAdapter(object):
    def __init__(self, url, key='', year=0, month=None,
                 from_date='2020-06-16', to_date='2020-06-30', username='', password=''):
        self.url = url or 'http://192.168.67.129:7777/redmine'
        if key == '':
            key = None
        self.key = key
        self.username = username
        self.password = password
        self.redmine = Redmine(url, key=key, username=username, password=password)
        self.current = self.redmine.user.get('current')
        if month is None:
            self.from_date = from_date
            self.to_date = to_date
        else:
            if not isinstance(month, int):
                raise ValueError('The month must be round')
            if year != 0 and not isinstance(year, int):
                raise ValueError('The year must be round')
            if month < 1 or month > 12:
                raise ValueError('The month must be between 1 and 12')
            first_day, last_day = self.get_month_first_day_and_last_day(year=year, month=month)
            self.from_date = str(first_day)
            self.to_date = str(last_day)

        if self.key is None:
            try:
                self.custom_session = self.create_custom_session()
            except Exception as e:
                self.custom_session = None
            if self.custom_session is None:
                click.echo('Warming: SPDM simulated user login failure, unable to use the project filtering function')
            else:
                click.echo('Info: SPDM simulates successful login of the user')
        else:
            self.custom_session = None
            click.echo('Warming: Current using token unable to use the project filtering function')

    def create_custom_session(self):
        login_url = '{0}/login'.format(self.redmine.url)
        session = requests.session()
        result = session.get(login_url)
        auth_data = {
            'username': self.username,
            'password': self.password,
        }
        if result.status_code == 200:
            login_html = etree.HTML(result.content, etree.HTMLParser())
            result = login_html.xpath('//input[@name="authenticity_token"]/./@value')
            if len(result) == 1:
                auth_data.update(authenticity_token=result[0])
                result = session.post(login_url, data=auth_data)
                if result.status_code == 200:
                    return session
        return None

    @staticmethod
    def get_month_first_day_and_last_day(year=None, month=None):
        """
        :param year: default current the year
        :param month:
        :return: first_day:
                 last_day:
        """
        if year:
            year = int(year)
        else:
            year = datetime.date.today().year

        if month:
            month = int(month)
        else:
            month = datetime.date.today().month

        first_day_week_day, month_range = calendar.monthrange(year, month)
        first_day = datetime.date(year=year, month=month, day=1)
        last_day = datetime.date(year=year, month=month, day=month_range)

        return first_day, last_day

    def get_work_times(self, offset, limit=20):
        work_times = self.redmine.time_entry.filter(offset=offset,
                                                    limit=limit,
                                                    from_date=self.from_date,
                                                    to_date=self.to_date)
        return work_times

    def get_project_by_identifier(self, identifier):
        """
        get project object
        :param identifier: id or string
        :return: project object
        """
        return self.redmine.project.get(identifier)

    def get_project_by_name(self, name):
        try:
            identifier = int(name)
        except ValueError:
            identifier = None
        if identifier is not None and isinstance(identifier, int):
            project = self.get_project_by_identifier(identifier)
        else:
            project = self.get_project_by_identifier(name)
        return project

    def _get_projects(self):
        stacks = [0]
        projects = Projects(self.redmine)
        work_times = []

        def loop():
            _work_times = self.get_work_times(0, limit=1)
            _work_times_num = len(_work_times)
            click.echo('Step one: Downloading data from SPDM,please waiting....')
            with click.progressbar(length=_work_times.total_count) as bar:
                while True:
                    offset = stacks.pop()
                    _work_times = self.get_work_times(offset)
                    for work_time in _work_times:
                        remote_user = work_time.user
                        remote_project = work_time.project
                        remote_issue = getattr(work_time, 'issue', None)
                        project = projects.get_project(remote_project)
                        user = project.get_user(remote_user)
                        if remote_issue is not None:
                            task = user.get_task(remote_issue)
                            work_time = task.get_work_time(work_time)
                        else:
                            logger.warning('object{0} has no attribute issue'.format(str(work_time)))
                        assert work_time is not None
                    work_times.extend(_work_times)
                    current_count = len(work_times)
                    bar.update(len(_work_times))
                    if current_count < _work_times.total_count:
                        stacks.append(current_count)
        try:
            loop()
        except IndexError:
            pass
        return projects

    def _get_sub_projects(self, project_id):
        sub_project_url = '{0}/projects/{1}/children'.format(self.redmine.url, project_id)
        try:
            response = self.custom_session.get(sub_project_url)
        except Exception as e:
            response = None
        if response and response.status_code == 200:
            items = json.loads(response.content)
            projects = items.get('children')
            return [CustomRemoteProject(project.get('id'), project.get('name')) for project in projects]
        return []

    def get_sub_projects(self, project):
        stack = [project]
        projects = []
        while True:
            try:
                current_project = stack.pop()
                if current_project not in projects:
                    yield current_project
                    projects.append(current_project)
            except IndexError:
                return
            stack.extend(self._get_sub_projects(current_project.id))

    def get_projects(self, redmine_project=None):
        projects = self._get_projects()
        if redmine_project is not None and self.custom_session is not None:
            projects = self.checkout_projects(projects, redmine_project)
        else:
            click.echo('Warming: Ignore project')
        return projects

    def checkout_projects(self, src_projects, redmine_project):
        try:
            entry_project = self.get_project_by_name(redmine_project)
        except Exception as e:
            entry_project = None
        if redmine_project is not None and entry_project is None:
            raise ValueError('The item named `{0}` could not be found'.format(redmine_project))
        click.echo('Locate the {0} project'.format(entry_project.name))
        sub_projects = self.get_sub_projects(entry_project)
        click.echo('Step two: Checkout projects,please waiting....')
        projects = []
        tests = []
        with click.progressbar(sub_projects) as bar:
            for dst_project in bar:
                tests.append(dst_project)
                project = src_projects.get_project_by_project_id(dst_project.id)
                if project is not None:
                    projects.append(project)
            src_projects.clear_resource()
            src_projects.extend_resource(projects)
        return src_projects

    def get_current_user_fullname(self):
        return '{0}{1}'.format(self.current.lastname, self.current.firstname)


def process(*args, **kwargs):
    enable_merge_cells = kwargs.pop('enable_merge_cells', True)
    redmine_project = kwargs.pop('project', None)

    redmine = RedmineAdapter(*args, **kwargs)
    projects = redmine.get_projects(redmine_project=redmine_project)

    adapter = ExcelAdapter("template.xlsx", "{0}--{1} created on {2}.xlsx".
                           format(redmine.from_date,
                                  redmine.to_date,
                                  datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')))
    work_table = WorkTable(adapter, projects, enable_merge=enable_merge_cells)
    work_table.process()


@click.command()
@click.option("--url", default='http://spdm/redmine/',
              help="server address example: http://spdm/redmine/")
@click.option("--key", default='', help="SPDM access token")
@click.option("--username", default='', help="SPDM username")
@click.option("--password", default='', help="SPDM password")
@click.option("--year", help="Statistical year,the default this year", default=0, type=click.IntRange(0, 9999))
@click.option("--month", help="Statistical month", required=True, type=click.IntRange(1, 12))
@click.option("--enable-merge-cells", default=False, help="enable merge cells", is_flag=True)
@click.option("--project", default=None, help="SPDM project identifier")
def gen_excel(url, key, year, month, username, password, enable_merge_cells, project):
    """Generate Excel"""
    try:
        process(url=url, key=key, year=year, month=month,
                username=username, password=password, enable_merge_cells=enable_merge_cells, project=project)
    except Exception as e:
        click.echo(str(e))


if __name__ == '__main__':
    gen_excel()
